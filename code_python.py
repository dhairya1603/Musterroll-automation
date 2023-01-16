import time
from openpyxl.reader import excel  
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By  
import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
My_path = r"C:\Users\Blast from Past\Downloads\Telegram Desktop\Book2.xlsx"  #excel file address
my_wb_obj = openpyxl.load_workbook(My_path)
my_sheet_obj = my_wb_obj.active
district_code= str('2726') #pleasen add a drop down menu later in future
district_name= str('UDAIPUR') #all caps
search_type= str('wrk') #usually "wrk" but must be checked on official website
state_code= str('RJ') #2 letter code in all caps
state_name= str("RAJASTHAN") #all caps
brouser = input('what brouser you use [F] for Firefox [E] for Edge : ')
if (brouser == 'F' or brouser == 'f' ):
     driver = webdriver.Firefox()
     driver.minimize_window()
elif (brouser == 'E' or brouser == 'e' ):
     driver = webdriver.Edge()
     driver.minimize_window()
else:
     print('please select one of the supppourted brouser')
print(" please go and collect today`s digest number from your selected brouser (with it attached a programme which doesn`t work)")
dig = input('enter today`s digest number: ') #visit site before on the given brouser
exel_st = 1 #you can begin by the number you left with in case of an error or perheps you can multithread
dha = 6 #number of codes present in excel
digest_number= str(dig) #last words in the end of a site (changes every 2400 hrs hence must be checked daily) 
print("programme starts")
#for websites in English
def englishsite():
 for i in range(exel_st+2,dha):
     my_cell_obj = my_sheet_obj.cell(row = i, column = 2)
     pg_number_xl=my_sheet_obj.cell(row = i, column = 6)
     mb_number_xl=my_sheet_obj.cell(row = i, column = 5)
     sc_date_xl=my_sheet_obj.cell(row = i, column = 4)
     sc_number_xl=my_sheet_obj.cell(row = i, column = 3)
     incode = str(my_cell_obj.value)
     print(incode)
     url = 'http://mnregaweb4.nic.in/netnrega/master_search1.aspx?flag=2&wsrch='+search_type+'&district_code='+district_code+'&state_name='+state_name+'&district_name='+district_name+'&short_name='+state_code+'&srch='+incode+'&Digest='+digest_number+''
     #url = 'http://mnregaweb4.nic.in/netnrega/master_search1.aspx?flag=2&wsrch=wrk&district_code=2728&state_name=RAJASTHAN&district_name=BANSWARA&short_name=RJ&srch=2728006248/LD/112908264006&Digest=Uzw01oWU5MCqnUzIfn/f2g'
     print(url)   
     driver.get(url)
     #time.sleep(20)
     driver.find_element_by_xpath('/html/body/form/div[3]/center[1]/div/table/tbody/tr[2]/td[4]/a').click()
     #time.sleep(10)
     try:
          sc_date= driver.find_element_by_xpath('/html/body/center/form/table[3]/tbody/tr[6]/td[1]/nobr/p/font[2]').text #or driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
          sc_number= driver.find_element_by_xpath('/html/body/center/form/table[3]/tbody/tr[6]/td[1]/nobr/p/font[1]').text #or driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
          sc_date_xl.value = str(sc_date)
          sc_number_xl.value = str(sc_number)
          my_wb_obj.save(My_path)
          print(sc_number)
          print(sc_date)
     except NoSuchElementException:
         pass
         #print("sanction not found")
         #sc_number= driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
         #sc_date= driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[2]').text
     try:
          sc_date= driver.find_element_by_xpath('/html/body/center/form/table[3]/tbody/tr[5]/td[1]/nobr/p/font[2]').text #or driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
          sc_number= driver.find_element_by_xpath('/html/body/center/form/table[3]/tbody/tr[5]/td[1]/nobr/p/font[1]').text #or driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
          sc_date_xl.value = str(sc_date)
          sc_number_xl.value = str(sc_number)
          my_wb_obj.save(My_path)
          print(sc_number)
          print(sc_date)
          print('section found')

     except NoSuchElementException:
         pass
         #print("sanction not found")
         #sc_number= driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
         #sc_date= driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[2]').text
     try:
          sc_date= driver.find_element_by_xpath('/html/body/center/form/table[3]/tbody/tr[7]/td[1]/nobr/p/font[2]').text #or driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
          sc_number= driver.find_element_by_xpath('/html/body/center/form/table[3]/tbody/tr[7]/td[1]/nobr/p/font[1]').text #or driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
          sc_date_xl.value = str(sc_date)
          sc_number_xl.value = str(sc_number)
          my_wb_obj.save(My_path)
          print(sc_number)
          print(sc_date)

     except NoSuchElementException:
         print("sanction not found")
         #sc_number= driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[1]').text
         #sc_date= driver.find_element_by_xpath('/html/body/center/form/table[2]/tbody/tr[8]/td[2]/font[2]').text

     try:
         all_children_by_css = driver.find_elements_by_xpath("/html/body/center/form/table[3]/tbody/tr[12]/td[2]/font/a")
         #print(all_children_by_css)
         n= len(all_children_by_css)
         y= str(n)
         z= "/html/body/center/form/table[3]/tbody/tr[12]/td[2]/font/a["+y+"]"
         driver.find_element_by_xpath(z).click()
         #time.sleep(5)
         pg_number= driver.find_element_by_id('ContentPlaceHolder1_page_no').text 
         mb_number= driver.find_element_by_id('ContentPlaceHolder1_mbno').text
         sc_number_xl.value = str(sc_number)
         sc_date_xl.value = str(sc_date)
         mb_number_xl.value = str(mb_number)
         pg_number_xl.value = str(pg_number)
         my_wb_obj.save(My_path)
         print(mb_number)
         print(pg_number)
         print(sc_date)
         print(sc_number)
         print("musterroll found")
     except NoSuchElementException:
         pass
         #print('no musterrolls')
     try:
         all_children_by_css = driver.find_elements_by_xpath("/html/body/center/form/table[3]/tbody/tr[11]/td[2]/font/a")
         #print(all_children_by_css)
         n= len(all_children_by_css)
         y= str(n)
         z= "/html/body/center/form/table[3]/tbody/tr[11]/td[2]/font/a["+y+"]"
         driver.find_element_by_xpath(z).click()
         #time.sleep(5)
         pg_number= driver.find_element_by_id('ContentPlaceHolder1_page_no').text 
         mb_number= driver.find_element_by_id('ContentPlaceHolder1_mbno').text
         sc_number_xl.value = str(sc_number)
         sc_date_xl.value = str(sc_date)
         mb_number_xl.value = str(mb_number)
         pg_number_xl.value = str(pg_number)
         my_wb_obj.save(My_path)
         print(mb_number)
         print(pg_number)
         print(sc_date)
         print(sc_number)
         print("musterroll found")
     except NoSuchElementException:
         pass
         #print('no musterrolls')
     try:
         all_children_by_css = driver.find_elements_by_xpath("/html/body/center/form/table[3]/tbody/tr[13]/td[2]/font/a")
         #print(all_children_by_css)
         n= len(all_children_by_css)
         y= str(n)
         z= "/html/body/center/form/table[3]/tbody/tr[13]/td[2]/font/a["+y+"]"
         driver.find_element_by_xpath(z).click()
         #time.sleep(5)
         pg_number= driver.find_element_by_id('ContentPlaceHolder1_page_no').text 
         mb_number= driver.find_element_by_id('ContentPlaceHolder1_mbno').text
         sc_number_xl.value = str(sc_number)
         sc_date_xl.value = str(sc_date)
         mb_number_xl.value = str(mb_number)
         pg_number_xl.value = str(pg_number)
         my_wb_obj.save(My_path)
         print(mb_number)
         print(pg_number)
         print(sc_date)
         print(sc_number)
         print("musterroll found")
     except NoSuchElementException:
         print('no musterrolls')
#code over         
def closed():
    driver.close() 
 
while True:
     englishsite()
     closed()       
